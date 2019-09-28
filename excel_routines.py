import openpyxl
import io
from contextlib import redirect_stdout


class Excel():
    def __init__(self, filename):
        self.filename = filename
        #open the default source xlsx file
        try:
            self.workbook = openpyxl.load_workbook(filename)
            self.sheet = self.workbook.active
            print('INFO: Excel file opened: ' + filename)
        except:
            print('ERROR: ' + filename + ' not found, please create a blank ' + filename + ' file')

        #find the row with no entry
        blank = False
        self.blankrow = 1

        while not blank:
            if not(self.sheet.cell(row = self.blankrow, column = 1).value):
                blank = True
            else:
                self.blankrow = self.blankrow + 1

    def write_last_row(self, *args):
        #args are the strings to be loaded in a single row
        try:
            i = 1
            for arg in args:
                self.sheet.cell(row = self.blankrow, column = i, value = str(arg))
                i = i + 1
            print(('INFO: Data write at row: ' + str(self.blankrow - 1)))
            
            self.blankrow = self.blankrow + 1
        except:
            print('INFO: excel write_last_row error')

    def write_last_row_list(self, list_input):
        #args are the strings to be loaded in a single row contained in a list
        for i in range(1, len(list_input) + 1): 
            # i is adjusted based on the excel file, list file adapts
            self.sheet.cell(row = self.blankrow, column = 1, value = str(list_input[i-1]))
            self.blankrow = self.blankrow + 1
        
        print(('INFO: Entry logged at row: ' + str(self.blankrow - 1)))
        
        

    def save(self, *args):
        if len(args) == 0:
            self.workbook.save(self.filename)
            print('INFO: Excel file saved: ' + self.filename)
        else:
            self.workbook.save(args[0])
            print('INFO: Excel file saved: ' + args[0])

    def read_column(self, column):
        column_list = []
        for row in range(1, self.sheet.max_row + 1):
            if self.sheet.cell(row = row, column = 1).value:
                column_list.append(self.sheet.cell(row = row, column = 1).value)

        return column_list

    def blank_file(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        workbook.save(filename)
        print(('INFO: Excel file created: ' + filename))



import openpyxl
from openpyxl.utils import get_column_letter

# This class works with openpyxl to add/delete rows and columns in Excel
# Openpyxl didn't have this feature to my knowledge, so I coded a way to do it
class Shift:
    def __init__(self, worksheet, selected_row_col):
        self.worksheet = worksheet
        self.selected = selected_row_col
        self.column = worksheet.max_column
        self.row = worksheet.max_row

    def insert_row(self): # Inserts row
        for i in range(self.row, self.selected-1, -1):
            for j in range(1, self.column+1):
                if not self.worksheet.cell(row=i, column=j).value:
                    self.worksheet[get_column_letter(j)+str(i+1)] = None
                else:
                    shift_here = self.worksheet.cell(row=i, column=j).value
                    self.worksheet[get_column_letter(j)+str(i+1)] = str(shift_here)
                if i == self.selected:
                    self.worksheet[get_column_letter(j)+str(i)] = None


    def insert_column(self): # Inserts column
        for i in range(self.column, self.selected-1, -1):
            for j in range(1, self.row+1):
                if not self.worksheet.cell(row=j, column=i).value:
                    self.worksheet[get_column_letter(i)+str(j)] = ''
                else:
                    shift_here = self.worksheet.cell(row=j, column=i).value
                    self.worksheet[get_column_letter(i+1)+str(j)] = str(shift_here)
                if i == self.selected:
                    self.worksheet[get_column_letter(i)+str(j)] = None


    def remove_row(self): # Deletes row
        for i in range(self.row, self.selected-1, -1):
            for j in range(1, self.column+1):
                if not self.worksheet.cell(row=i+1, column=j).value:
                    self.worksheet[get_column_letter(j)+str(i)] = None
                else:
                    shift_here = self.worksheet.cell(row=i+1, column=j).value
                    self.worksheet[get_column_letter(j)+str(i)] = str(shift_here)
                if i == self.row-1:
                    self.worksheet[get_column_letter(j)+str(i+1)] = None


    def remove_column(self): # Deletes column
        for i in range(self.column-1, self.selected-1, -1):
            for j in range(1, self.row+1):
                if not self.worksheet.cell(row=j, column=i+1).value:
                    self.worksheet[get_column_letter(i)+str(j)] = None
                else:
                    shift_here = self.worksheet.cell(row=j, column=i+1).value
                    self.worksheet[get_column_letter(i)+str(j)] = str(shift_here)
                if i == self.column-1:
                    self.worksheet[get_column_letter(i+1)+str(j)] = None